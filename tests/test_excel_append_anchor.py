import os
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook

from hh_research.client import VacancyRef
from hh_research.pipeline import append_refs_to_template_file


class TestExcelAppendAnchor(unittest.TestCase):
    def _make_template_xlsx(self, sheet_name: str, col_title: int, col_link: int) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Simulate a partially filled template: only title/link are filled on start_row.
        start_row = 2
        ws.cell(1, col_title, "Header Title")
        ws.cell(1, col_link, "Header Link")
        ws.cell(start_row, col_title, "Some Vacancy Title")
        ws.cell(start_row, col_link, "https://hh.ru/vacancy/123")

        tf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tf.close()
        wb.save(tf.name)
        return tf.name

    def test_next_row_uses_title_and_link(self) -> None:
        sheet_name = "Sheet1"
        col_title = 1
        col_keywords = 2
        col_skills = 3
        col_id = 4
        col_link = 5
        start_row = 2

        template_path = self._make_template_xlsx(sheet_name, col_title=col_title, col_link=col_link)
        out_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        out_path.close()

        try:
            refs = [VacancyRef(vacancy_id="123")]
            with patch("hh_research.pipeline.run_export_on_worksheet") as mock_run:
                # processed, next_row_after_last_block, errors, summary
                mock_run.return_value = (0, 10, [], {"coverage": {}})

                processed, next_row_after, errors, summary = append_refs_to_template_file(
                    excel_path=template_path,
                    sheet_name=sheet_name,
                    refs=refs,
                    token=None,
                    start_row=start_row,
                    col_title=col_title,
                    col_keywords=col_keywords,
                    col_skills=col_skills,
                    col_id=col_id,
                    col_link=col_link,
                    kw_top_n=30,
                    kw_max_ngram=3,
                    sleep_s=0,
                    out_path=out_path.name,
                )

                # With title/link filled on row 2, next_row must be row 3.
                self.assertEqual(mock_run.call_args.kwargs["next_row"], 3)
                self.assertEqual(processed, 0)
                self.assertEqual(errors, [])
                self.assertEqual(summary.get("coverage", {}), {})
                self.assertEqual(next_row_after, 10)
        finally:
            if os.path.exists(template_path):
                os.unlink(template_path)
            if os.path.exists(out_path.name):
                os.unlink(out_path.name)


if __name__ == "__main__":
    unittest.main()

