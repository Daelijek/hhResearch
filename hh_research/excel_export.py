from typing import Tuple, List, Optional

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from hh_research.keywords import extract_keywords_simple, extract_text_from_html

# Approximate column widths (openpyxl units ≈ character count)
DEFAULT_COL_WIDTHS: dict[int, float] = {
    1: 42.0,  # title
    2: 32.0,  # keywords
    3: 32.0,  # skills
    4: 14.0,  # id
    5: 48.0,  # link
    6: 32.0,  # unique keywords
    7: 32.0,  # unique skills
}

# Header: bold 12pt, each column own background (Excel-style palette, white text)
HEADER_FILLS: dict[int, str] = {
    1: "4472C4",  # blue
    2: "ED7D31",  # orange
    3: "70AD47",  # green
    4: "7030A0",  # purple
    5: "00B0F0",  # cyan
    6: "A5A5A5",  # gray (unique keywords)
    7: "FFC000",  # amber (unique skills)
}

DUPLICATE_FILL_HEX = "C6EFCE"  # light green (Excel-style) for repeated values


def extract_skills_and_keywords(
    vac: dict,
    kw_top_n: int,
    kw_max_ngram: int,
) -> Tuple[str, str, str, List[str], List[str]]:
    vid = str(vac.get("id") or "").strip()
    title = str(vac.get("name") or "").strip()
    link = vac.get("alternate_url") or vac.get("url") or ""
    link = str(link).strip()
    if not link and vid:
        link = f"https://hh.ru/vacancy/{vid}"

    key_skills = vac.get("key_skills") or []
    skill_names: list[str] = []
    for s in key_skills:
        name = None
        if isinstance(s, dict):
            name = s.get("name")
        elif isinstance(s, str):
            name = s
        if name and str(name).strip():
            skill_names.append(str(name).strip())

    description_html = vac.get("description") or ""
    description_text = extract_text_from_html(description_html)
    keywords = extract_keywords_simple(
        description_text,
        top_n=kw_top_n,
        max_ngram=kw_max_ngram,
    )
    keywords = [str(k).strip() for k in keywords if k and str(k).strip()]

    return title, vid, link, keywords, skill_names


def extract_vacancy_identity(vac: dict) -> Tuple[str, str, str]:
    """Fast extraction of fields used for Excel writing (no HTML parsing)."""
    vid = str(vac.get("id") or "").strip()
    title = str(vac.get("name") or "").strip()
    link = vac.get("alternate_url") or vac.get("url") or ""
    link = str(link).strip()
    if not link and vid:
        link = f"https://hh.ru/vacancy/{vid}"
    return title, vid, link


def create_export_workbook(sheet_name: str = "Sheet1") -> Tuple[Workbook, Worksheet]:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.cell(1, 1, "Vacancy Title")
    ws.cell(1, 2, "Key Words")
    ws.cell(1, 3, "Key Skills (...)")
    ws.cell(1, 4, "ID")
    ws.cell(1, 5, "Link")
    ws.cell(1, 6, "Unique Keywords")
    ws.cell(1, 7, "Unique Skills")
    return wb, ws


def style_export_column_widths(
    ws: Worksheet,
    col_title: int,
    col_keywords: int,
    col_skills: int,
    col_id: int,
    col_link: int,
    col_unique_keywords: int,
    col_unique_skills: int,
) -> None:
    mapping = {
        col_title: DEFAULT_COL_WIDTHS[1],
        col_keywords: DEFAULT_COL_WIDTHS[2],
        col_skills: DEFAULT_COL_WIDTHS[3],
        col_id: DEFAULT_COL_WIDTHS[4],
        col_link: DEFAULT_COL_WIDTHS[5],
        col_unique_keywords: DEFAULT_COL_WIDTHS[6],
        col_unique_skills: DEFAULT_COL_WIDTHS[7],
    }
    for col_idx, w in mapping.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def style_export_header_row(
    ws: Worksheet,
    header_row: int,
    col_title: int,
    col_keywords: int,
    col_skills: int,
    col_id: int,
    col_link: int,
    col_unique_keywords: int,
    col_unique_skills: int,
) -> None:
    # Relative ordering matters for HEADER_FILLS palette.
    cols = (
        col_title,
        col_keywords,
        col_skills,
        col_id,
        col_link,
        col_unique_keywords,
        col_unique_skills,
    )
    font = Font(bold=True, size=12, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(header_row, c)
        fill_hex = HEADER_FILLS.get(i, "4472C4")
        cell.font = font
        cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        cell.alignment = align
    ws.row_dimensions[header_row].height = 22


def _first_data_row(ws: Worksheet, header_row: int, cols: Tuple[int, ...]) -> int:
    """Row index of first cell with content below header (for CF range)."""
    last = ws.max_row or header_row
    for r in range(header_row + 1, last + 1):
        for c in cols:
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                return r
    return header_row + 1


def apply_duplicate_cell_highlight(
    ws: Worksheet,
    col_idx: int,
    start_row: int,
    end_row: int,
) -> None:
    if end_row < start_row:
        return
    col = get_column_letter(col_idx)
    cell_range = f"{col}{start_row}:{col}{end_row}"
    fill = PatternFill(
        start_color=DUPLICATE_FILL_HEX,
        end_color=DUPLICATE_FILL_HEX,
        fill_type="solid",
    )
    # Highlight cell if the same value appears more than once in this column (non-empty)
    formula = f'AND({col}{start_row}<>"",COUNTIF(${col}${start_row}:${col}${end_row},{col}{start_row})>1)'
    rule = FormulaRule(formula=[formula], fill=fill)
    ws.conditional_formatting.add(cell_range, rule)


def finalize_export_sheet(
    ws: Worksheet,
    col_title: int,
    col_keywords: int,
    col_skills: int,
    col_id: int,
    col_link: int,
    col_unique_keywords: int,
    col_unique_skills: int,
    *,
    header_row: int = 1,
    data_start_row: Optional[int] = None,
    style_headers: bool = True,
) -> None:
    style_export_column_widths(
        ws,
        col_title,
        col_keywords,
        col_skills,
        col_id,
        col_link,
        col_unique_keywords,
        col_unique_skills,
    )
    if style_headers:
        style_export_header_row(
            ws,
            header_row,
            col_title,
            col_keywords,
            col_skills,
            col_id,
            col_link,
            col_unique_keywords,
            col_unique_skills,
        )
    cols_scan = (col_title, col_keywords, col_skills, col_id, col_link, col_unique_keywords, col_unique_skills)
    data_row = (
        data_start_row
        if data_start_row is not None
        else _first_data_row(ws, header_row, cols_scan)
    )
    max_r = ws.max_row or data_row
    if max_r >= data_row:
        apply_duplicate_cell_highlight(ws, col_keywords, data_row, max_r)
        apply_duplicate_cell_highlight(ws, col_skills, data_row, max_r)
    lo = min(cols_scan)
    hi = max(cols_scan)
    ws.freeze_panes = ws.cell(data_row, lo).coordinate
    for row in ws.iter_rows(min_row=data_row, max_row=max_r, min_col=lo, max_col=hi):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def find_next_row_after_max(ws: Worksheet, col: int, start_row: int = 2) -> int:
    last = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is not None and str(v).strip() != "":
            last = r
    return last + 1


def write_vacancy_row_aligned(
    ws: Worksheet,
    vac: dict,
    start_row: int,
    col_title: int,
    col_keywords: int,
    col_skills: int,
    col_id: int,
    col_link: int,
    kw_top_n: int,
    kw_max_ngram: int,
    *,
    keywords: List[str] | None = None,
    skill_names: List[str] | None = None,
) -> int:
    title, vid, link = extract_vacancy_identity(vac)

    # Pipeline передает keywords/skills уже вычисленными; это важно для скорости
    # и для того, чтобы эвристика извлечения выполнялась один раз на вакансию.
    if keywords is None or skill_names is None:
        _, _, _, computed_keywords, computed_skills = extract_skills_and_keywords(
            vac,
            kw_top_n=kw_top_n,
            kw_max_ngram=kw_max_ngram,
        )
        if keywords is None:
            keywords = computed_keywords
        if skill_names is None:
            skill_names = computed_skills

    keywords = keywords or []
    skill_names = skill_names or []

    max_len = max(len(keywords), len(skill_names), 1)

    for i in range(max_len):
        row = start_row + i

        if i == 0:
            if title:
                ws.cell(row, col_title, title)
            if vid:
                ws.cell(row, col_id, vid)
            if link:
                ws.cell(row, col_link, link)

        if i < len(keywords):
            ws.cell(row, col_keywords, keywords[i])
        if i < len(skill_names):
            ws.cell(row, col_skills, skill_names[i])

    return start_row + max_len
