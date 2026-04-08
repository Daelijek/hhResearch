import io
import os
import time
from datetime import datetime
from typing import Callable, List, Optional, Tuple, Dict, Any

import requests
from openpyxl import load_workbook
from collections import Counter

from hh_research.client import VacancyRef, fetch_vacancy, search_vacancies
from hh_research.excel_export import (
    create_export_workbook,
    finalize_export_sheet,
    find_next_row_after_max,
    extract_skills_and_keywords,
    write_vacancy_row_aligned,
)


def build_default_out_path(
    project_dir: str,
    out_dir: str,
    mode: str,
    out_prefix: str,
    template_basename: str,
) -> str:
    os.makedirs(os.path.join(project_dir, out_dir), exist_ok=True)
    _, ext = os.path.splitext(template_basename)
    if not ext:
        ext = ".xlsx"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_name = f"{out_prefix}_{mode}_filled_{timestamp}{ext}"
    return os.path.join(project_dir, out_dir, out_name)


def collect_refs_auto(
    session: requests.Session,
    token: Optional[str],
    queries: List[str],
    pages: int,
    per_page: int,
    dedupe_vacancies: bool,
    search_sleep_s: float,
    *,
    employer_id: Optional[str] = None,
    area: Optional[str] = None,
    experience: Optional[str] = None,
    period: Optional[int] = None,
) -> Tuple[List[VacancyRef], int]:
    """Returns (unique_refs, raw_id_hits) where raw_id_hits counts every search item with a valid id (incl. duplicates)."""
    seen_ids: set[str] = set()
    vacancy_ids_to_process: List[VacancyRef] = []
    raw_id_hits = 0
    for q in queries:
        for page in range(pages):
            items = search_vacancies(
                session,
                token,
                q,
                page=page,
                per_page=per_page,
                employer_id=employer_id,
                area=area,
                experience=experience,
                period=period,
            )
            for it in items:
                vid = str(it.get("id") or "").strip()
                if not vid:
                    continue
                raw_id_hits += 1
                if dedupe_vacancies:
                    if vid in seen_ids:
                        continue
                    seen_ids.add(vid)
                vacancy_ids_to_process.append(VacancyRef(vacancy_id=vid))
            if search_sleep_s > 0:
                time.sleep(search_sleep_s)
    return vacancy_ids_to_process, raw_id_hits


def compute_summary_for_refs(
    refs: List[VacancyRef],
    session: requests.Session,
    token: Optional[str],
    *,
    kw_top_n: int,
    kw_max_ngram: int,
    sleep_s: float,
    on_progress: Optional[Callable[[int, int], None]] = None,
) -> Tuple[List[str], Dict[str, Any]]:
    total = len(refs)
    processed = 0
    attempted = 0
    errors: List[str] = []
    keyword_freq: Counter[str] = Counter()
    skill_freq: Counter[str] = Counter()
    with_key_skills = 0
    without_key_skills = 0
    without_description = 0

    if on_progress and total > 0:
        on_progress(0, total)

    for ref in refs:
        vid = ref.vacancy_id
        try:
            vac = fetch_vacancy(session, token, vid)
        except Exception as e:
            errors.append(f"{vid}: {e}")
            attempted += 1
            if on_progress and total > 0:
                on_progress(attempted, total)
            continue

        _title, _vid, _link, keywords, skill_names = extract_skills_and_keywords(
            vac,
            kw_top_n=kw_top_n,
            kw_max_ngram=kw_max_ngram,
        )
        for s in skill_names:
            if s:
                skill_freq[s] += 1
        for k in keywords:
            if k:
                keyword_freq[k] += 1

        if skill_names:
            with_key_skills += 1
        else:
            without_key_skills += 1
        if not (vac.get("description") or "").strip():
            without_description += 1

        processed += 1
        attempted += 1
        if on_progress and total > 0:
            # Keep it smooth for small totals; throttle for larger sets.
            if attempted == total:
                on_progress(attempted, total)
            elif total <= 20:
                on_progress(attempted, total)
            elif total <= 100 and attempted % 5 == 0:
                on_progress(attempted, total)
            elif attempted % 10 == 0:
                on_progress(attempted, total)
        if sleep_s > 0:
            time.sleep(sleep_s)

    reason_counter: Counter[str] = Counter()
    for msg in errors:
        reason = msg.split(": ", 1)[1].strip() if ": " in msg else msg.strip()
        if len(reason) > 220:
            reason = reason[:217] + "..."
        reason_counter[reason] += 1

    successful = processed
    summary: Dict[str, Any] = {
        "requested": total,
        "processed": processed,
        "errors": len(errors),
        "top_skills": [{"name": k, "count": v} for k, v in skill_freq.most_common(20)],
        "top_keywords": [{"name": k, "count": v} for k, v in keyword_freq.most_common(20)],
        "coverage": {
            "successful": successful,
            "with_key_skills": with_key_skills,
            "without_key_skills": without_key_skills,
            "without_description": without_description,
            "key_skills_rate": round(with_key_skills / successful, 4) if successful else 0.0,
        },
        "error_breakdown": [{"reason": r, "count": c} for r, c in reason_counter.most_common(10)],
    }
    return errors, summary


def run_export_on_worksheet(
    ws,
    refs: List[VacancyRef],
    session: requests.Session,
    token: Optional[str],
    next_row: int,
    col_title: int,
    col_keywords: int,
    col_skills: int,
    col_id: int,
    col_link: int,
    kw_top_n: int,
    kw_max_ngram: int,
    sleep_s: float,
    on_progress: Optional[Callable[[int, int], None]] = None,
    *,
    format_headers: bool = True,
) -> Tuple[int, int, List[str], Dict[str, Any]]:
    """
    Fetches each vacancy and appends blocks to ws starting at next_row.
    Returns (processed_count, next_row_after_last_block, list of error messages per failed id).
    """
    total = len(refs)
    processed = 0
    errors: List[str] = []
    row = next_row
    skill_freq: Counter[str] = Counter()
    keyword_freq: Counter[str] = Counter()
    with_key_skills = 0
    without_key_skills = 0
    without_description = 0

    if on_progress and total > 0:
        on_progress(0, total)

    for ref in refs:
        vid = ref.vacancy_id
        try:
            vac = fetch_vacancy(session, token, vid)
        except Exception as e:
            errors.append(f"{vid}: {e}")
            continue
        _title, _vid, _link, keywords, skill_names = extract_skills_and_keywords(
            vac,
            kw_top_n=kw_top_n,
            kw_max_ngram=kw_max_ngram,
        )
        for s in skill_names:
            if s:
                skill_freq[s] += 1
        for k in keywords:
            if k:
                keyword_freq[k] += 1

        if skill_names:
            with_key_skills += 1
        else:
            without_key_skills += 1
        if not (vac.get("description") or "").strip():
            without_description += 1

        row = write_vacancy_row_aligned(
            ws=ws,
            vac=vac,
            start_row=row,
            col_title=col_title,
            col_keywords=col_keywords,
            col_skills=col_skills,
            col_id=col_id,
            col_link=col_link,
            kw_top_n=kw_top_n,
            kw_max_ngram=kw_max_ngram,
            keywords=keywords,
            skill_names=skill_names,
        )
        processed += 1
        if sleep_s > 0:
            time.sleep(sleep_s)
        if on_progress and total > 0:
            # Smooth for small totals; throttle for large ones to avoid excessive status writes.
            if processed == total:
                on_progress(processed, total)
            elif total <= 20:
                on_progress(processed, total)
            elif total <= 100 and processed % 5 == 0:
                on_progress(processed, total)
            elif processed % 10 == 0:
                on_progress(processed, total)

    # Unique columns are placed right after "Link" by default.
    col_unique_keywords = col_link + 1
    col_unique_skills = col_link + 2

    # Ensure headers are present even when we don't restyle (CLI append uses format_headers=False).
    ws.cell(1, col_unique_keywords, "Unique Keywords")
    ws.cell(1, col_unique_skills, "Unique Skills")

    # Popularity = frequency across the whole sheet (including already existing template content).
    # We reconstruct it from filled values in Key Words / Key Skills columns.
    keyword_freq_all: Counter[str] = Counter()
    skill_freq_all: Counter[str] = Counter()
    last_row = ws.max_row or 1
    for r in range(2, last_row + 1):
        kw_val = ws.cell(r, col_keywords).value
        if kw_val is not None:
            kw_str = str(kw_val).strip()
            if kw_str:
                keyword_freq_all[kw_str] += 1
        sk_val = ws.cell(r, col_skills).value
        if sk_val is not None:
            sk_str = str(sk_val).strip()
            if sk_str:
                skill_freq_all[sk_str] += 1

    # Clear previous unique lists (if template already had them) and rewrite fresh.
    for r in range(2, last_row + 1):
        ws.cell(r, col_unique_keywords).value = None
        ws.cell(r, col_unique_skills).value = None

    # Write unique values sorted by (-count) and then by text for deterministic ordering.
    sorted_keywords = sorted(keyword_freq_all.items(), key=lambda x: (-x[1], x[0]))
    sorted_skills = sorted(skill_freq_all.items(), key=lambda x: (-x[1], x[0]))

    for i, (name, _count) in enumerate(sorted_keywords):
        ws.cell(2 + i, col_unique_keywords, name)
    for i, (name, _count) in enumerate(sorted_skills):
        ws.cell(2 + i, col_unique_skills, name)

    # Update top-* summary to match unique columns.
    skill_freq = skill_freq_all
    keyword_freq = keyword_freq_all

    finalize_export_sheet(
        ws=ws,
        col_title=col_title,
        col_keywords=col_keywords,
        col_skills=col_skills,
        col_id=col_id,
        col_link=col_link,
        col_unique_keywords=col_unique_keywords,
        col_unique_skills=col_unique_skills,
        header_row=1,
        data_start_row=None,
        style_headers=format_headers,
    )

    reason_counter: Counter[str] = Counter()
    for msg in errors:
        reason = msg.split(": ", 1)[1].strip() if ": " in msg else msg.strip()
        if len(reason) > 220:
            reason = reason[:217] + "..."
        reason_counter[reason] += 1

    successful = processed
    summary: Dict[str, Any] = {
        "requested": total,
        "processed": processed,
        "errors": len(errors),
        "top_skills": [{"name": k, "count": v} for k, v in skill_freq.most_common(20)],
        "top_keywords": [{"name": k, "count": v} for k, v in keyword_freq.most_common(20)],
        "coverage": {
            "successful": successful,
            "with_key_skills": with_key_skills,
            "without_key_skills": without_key_skills,
            "without_description": without_description,
            "key_skills_rate": round(with_key_skills / successful, 4) if successful else 0.0,
        },
        "error_breakdown": [
            {"reason": r, "count": c} for r, c in reason_counter.most_common(10)
        ],
    }

    return processed, row, errors, summary


def export_refs_to_xlsx_bytes(
    refs: List[VacancyRef],
    token: Optional[str],
    kw_top_n: int = 30,
    kw_max_ngram: int = 3,
    sleep_s: float = 0.2,
    col_title: int = 1,
    col_keywords: int = 2,
    col_skills: int = 3,
    col_id: int = 4,
    col_link: int = 5,
    start_row: int = 2,
    sheet_name: str = "Sheet1",
    on_progress: Optional[Callable[[int, int], None]] = None,
) -> Tuple[bytes, int, List[str], Dict[str, Any]]:
    session = requests.Session()
    wb, ws = create_export_workbook(sheet_name=sheet_name)
    processed, _, errors, summary = run_export_on_worksheet(
        ws=ws,
        refs=refs,
        session=session,
        token=token,
        next_row=start_row,
        col_title=col_title,
        col_keywords=col_keywords,
        col_skills=col_skills,
        col_id=col_id,
        col_link=col_link,
        kw_top_n=kw_top_n,
        kw_max_ngram=kw_max_ngram,
        sleep_s=sleep_s,
        on_progress=on_progress,
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), processed, errors, summary


def append_refs_to_template_file(
    excel_path: str,
    sheet_name: str,
    refs: List[VacancyRef],
    token: Optional[str],
    start_row: int,
    col_title: int,
    col_keywords: int,
    col_skills: int,
    col_id: int,
    col_link: int,
    kw_top_n: int,
    kw_max_ngram: int,
    sleep_s: float,
    out_path: str,
    on_progress: Optional[Callable[[int, int], None]] = None,
) -> Tuple[int, int, List[str], Dict[str, Any]]:
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    next_row = max(
        find_next_row_after_max(ws, col_title, start_row),
        find_next_row_after_max(ws, col_keywords, start_row),
        find_next_row_after_max(ws, col_skills, start_row),
        find_next_row_after_max(ws, col_id, start_row),
        find_next_row_after_max(ws, col_link, start_row),
    )
    session = requests.Session()
    processed, next_row_after, errors, summary = run_export_on_worksheet(
        ws=ws,
        refs=refs,
        session=session,
        token=token,
        next_row=next_row,
        col_title=col_title,
        col_keywords=col_keywords,
        col_skills=col_skills,
        col_id=col_id,
        col_link=col_link,
        kw_top_n=kw_top_n,
        kw_max_ngram=kw_max_ngram,
        sleep_s=sleep_s,
        on_progress=on_progress,
        format_headers=False,
    )
    wb.save(out_path)
    return processed, next_row_after, errors, summary
